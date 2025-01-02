import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

def format_excel_file(input_file, output_file):
    # Load the unformatted Excel file
    #input_file = "02_unformatted.xlsx"
    #output_file = "NEW_FORMATTED.xlsx"

    wb = openpyxl.load_workbook(input_file)
    sheet = wb.active

    # Define a bold black border style
    bold_black_border = Border(bottom=Side(style="thick", color="000000"))

    # Iterate over the rows, starting from the second row (assuming the first row is the header)
    previous_id = None
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), start=2):
        current_id = row[0].value  # Assuming the ID column is the first column (A)

        # Check if the current ID is different from the previous ID
        if current_id != previous_id and previous_id is not None:
            # Apply the bold black border to the row before the current one
            for col in range(1, sheet.max_column + 1):
                sheet.cell(row=row_index - 1, column=col).border = bold_black_border

        # Update the previous_id for the next iteration
        previous_id = current_id
    
    # Center column "ANZAHL" (Assumed as Column B)
    for row in range(1, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=2)  # Access the second column (B)
        cell.alignment = Alignment(horizontal="center")

    # Filter "SKU" entries with "4S" and format as bold
    for row in range(1, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=3)  # Access Column C
        if cell.value and "4S" in str(cell.value):
            cell.font = Font(bold=True)

    # Filter "SKU" entries with "6S" and format as italics and underlined
    for row in range(1, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=3)  # Access Column C
        if cell.value and "6S" in str(cell.value):
            cell.font = Font(italic=True, underline="single")

    # Filter "SKU" entries with "-2-" and format as bold
    for row in range(1, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=3)  # Access Column C
        if cell.value and "-2-" in str(cell.value):
            cell.font = Font(bold=True)

    # Highlight rows with "ANZAHL" > 1 in light blue
    for row in range(2, sheet.max_row + 1):  # Skip header row
        quantity_cell = sheet.cell(row=row, column=2)  # Column E
        if quantity_cell.value and quantity_cell.value > 1:
            fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            for col in range(1, sheet.max_column + 1):
                sheet.cell(row=row, column=col).fill = fill

    # Determine the maximum width needed for column C
    max_width = 0
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=3, max_col=3):
        cell_value = row[0].value  # Access the cell in column C
        if cell_value is not None:
            max_width = max(max_width, len(str(cell_value)))

    # Adjust the column width
    sheet.column_dimensions['C'].width = max_width

    #Show only the first unique ID in the ID column
    # Track seen IDs
    seen_ids = set()

    # Iterate over the rows starting from the second row (assuming the first row is the header)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        cell_id = row[0]  # Assuming the ID column is the first column (A)
        cell_id.alignment = Alignment(horizontal="center")

        # Check if the ID has been seen before
        if cell_id.value in seen_ids:
            cell_id.value = None  # Clear the duplicate ID
        else:
            seen_ids.add(cell_id.value)  # Add the new ID to the set

    # Add a filter to the sheet
    sheet.auto_filter.ref = sheet.dimensions

    # Center-align the header row (assumed to be the first row)
    for cell in sheet[1]:  # sheet[1] refers to the first row
        cell.alignment = Alignment(horizontal="center")

    # Save the formatted file
    wb.save(output_file)

